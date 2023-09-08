import openpyxl


def rowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def readData(file, sheetname, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rowno, column=colno).value

def writeData(file, sheetname, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rowno, column=colno).value = data
    workbook.save(file)

# import openpyxl
#
#
# def readData(file, sheetname, rowno, colno):
#     workbook = openpyxl.load_workbook(sheetname)
#     sheet = workbook[file]
#     return sheet.cell(row=rowno, column=colno).value
#
#
# def writeData(file, sheetname, rowno, colno, data):
#     workbook = openpyxl.load_workbook(sheetname)
#     sheet = workbook[file]
#     sheet.cell(row=rowno, column=colno).value = data
#
#
# def rowCount(file, sheetname):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetname]
#     return sheet.max_row
